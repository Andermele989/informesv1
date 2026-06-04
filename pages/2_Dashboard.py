import os
import sys
import re
import io
import datetime

import streamlit as st
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from dotenv import load_dotenv



from database import SessionLocal
import models

load_dotenv(override=True)

OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-3.5-flash")
PLOT_CONFIG = {
    "displayModeBar": False,
    "responsive": True,
    "staticPlot": False,
}
PREMIUM_COLORS = [
    "#14f1b2", "#38bdf8", "#a78bfa", "#fbbf24", "#fb7185",
    "#22d3ee", "#34d399", "#818cf8", "#f472b6", "#2dd4bf",
    "#c084fc", "#67e8f9", "#a3e635", "#f0abfc", "#fdba74",
]

try:
    import google.generativeai as genai
    from openai import OpenAI
except ImportError as e:
    st.warning(f"IA no disponible: {e}")

if not st.session_state.get("logged_in", False):
    st.warning("Por favor, inicia sesión.")
    st.stop()

# ========================= CARGAR ESTILOS RESPONSIVOS =========================
from styles_loader import cargar_estilos_responsivos

cargar_estilos_responsivos()

st.markdown("""
<div class="dash-header">
    <h1>📈 Dashboard de Rendimiento</h1>
    <p>Análisis estadístico y proyecciones del grupo</p>
</div>
""", unsafe_allow_html=True)

# ========================= DATA =========================
db = SessionLocal()
try:
    reports = db.query(models.MonthlyReport).all()
    users = db.query(models.User).all()
    publishers = db.query(models.Publisher).all()
    groups = db.query(models.Group).all()
finally:
    db.close()

if not reports:
    st.info("📭 No hay informes registrados aún. Ve a 'Registro' para añadir el primer informe.")
    st.stop()

publisher_map = {p.id: p.name for p in publishers}
publisher_group_map = {p.id: p.group_id for p in publishers}
group_map = {g.id: g.name for g in groups}

data = []
for r in reports:
    pub_name = r.full_name or publisher_map.get(r.publisher_id, "Desconocido")
    grp_name = group_map.get(publisher_group_map.get(r.publisher_id), "Sin Grupo")
    data.append({
        "id": r.id,
        "publisher_id": r.publisher_id,
        "Publicador": pub_name,
        "Grupo": grp_name,
        "Mes": r.month,
        "Privilegios": r.assigned_privileges,
        "Cursos Bíblicos": r.bible_courses,
        "Informe": r.service_report,
        "Notas": r.notes or ""
    })

# Auto-añadir inactivos
meses_registrados = sorted(list(set([r.month for r in reports])), reverse=True)
publishers_inactivos = [p for p in publishers if p.is_inactive]
for p in publishers_inactivos:
    grp_name = group_map.get(p.group_id, "Sin Grupo")
    for m in meses_registrados:
        if not any(d["Publicador"] == p.name and d["Mes"] == m for d in data):
            data.append({
                "id": None,
                "publisher_id": p.id,
                "Publicador": p.name,
                "Grupo": grp_name,
                "Mes": m,
                "Privilegios": "Inactivo",
                "Cursos Bíblicos": 0,
                "Informe": "INACTIVO (No predicó)",
                "Notas": "Registrado automático por el sistema"
            })

df = pd.DataFrame(data)

# ========================= FUNCIONES =========================
def extraer_horas(texto):
    if not isinstance(texto, str): return 0
    matches = re.findall(r'(\d+)\s*(?:horas|hrs|h)', texto.lower())
    return sum(int(m) for m in matches) if matches else 0

def obtener_año_servicio(mes_str):
    try:
        y, m = map(int, mes_str.split('-'))
        inicio = y if m >= 9 else y - 1
        return f"{inicio}-{inicio + 1}"
    except:
        return "Desconocido"

def nombre_corto(nombre, limite=18):
    nombre = str(nombre or "")
    return nombre if len(nombre) <= limite else f"{nombre[:limite - 2]}.."

def aplicar_layout_premium(fig, height=380, margin=None, mobile=False):
    fig.update_layout(
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Outfit, Inter, sans-serif", size=12 if mobile else 13, color="#f8fafc"),
        margin=margin or dict(t=28, b=54, l=48, r=28),
        height=height,
        hoverlabel=dict(
            bgcolor="#111827",
            bordercolor="rgba(20,241,178,0.35)",
            font=dict(color="#f8fafc", family="Outfit")
        ),
        transition=dict(duration=650, easing="cubic-in-out")
    )
    fig.update_xaxes(
        showgrid=False,
        tickfont=dict(size=10 if mobile else 12, color="#cbd5e1", family="Outfit"),
        zeroline=False,
    )
    fig.update_yaxes(
        showgrid=True,
        gridcolor="rgba(148,163,184,0.10)",
        tickfont=dict(size=10 if mobile else 12, color="#94a3b8", family="Outfit"),
        zeroline=False,
    )
    return fig

df["Año Servicio"] = df["Mes"].apply(obtener_año_servicio)
df["Horas"] = df["Informe"].apply(extraer_horas)

# ========================= FILTROS =========================
st.sidebar.header("🔍 Filtros de Informe")

# --- Año de servicio ---
años_servicio_disp = sorted(df["Año Servicio"].unique().tolist(), reverse=True)
año_servicio_actual = obtener_año_servicio(datetime.datetime.now().strftime("%Y-%m"))
default_index_año = años_servicio_disp.index(año_servicio_actual) if año_servicio_actual in años_servicio_disp else 0

año_servicio_sel = st.sidebar.selectbox(
    "📅 Año de Servicio",
    ["Todos"] + años_servicio_disp,
    index=default_index_año + 1
)

df_filtrado_año = df if año_servicio_sel == "Todos" else df[df["Año Servicio"] == año_servicio_sel]

# --- Meses ---
meses_disponibles = sorted(df_filtrado_año["Mes"].unique().tolist(), reverse=True)
mes_actual = datetime.datetime.now().strftime("%Y-%m")

meses_seleccionados = st.sidebar.multiselect(
    "📆 Mes(es)",
    meses_disponibles,
    default=[mes_actual] if mes_actual in meses_disponibles else ([meses_disponibles[0]] if meses_disponibles else [])
)

if meses_seleccionados:
    df_filtrado_mes = df_filtrado_año[df_filtrado_año["Mes"].isin(meses_seleccionados)]
else:
    df_filtrado_mes = df_filtrado_año

# --- Grupo ---
st.sidebar.divider()
grupos_disponibles = sorted(df_filtrado_mes["Grupo"].unique().tolist())
grupo_sel = st.sidebar.selectbox(
    "🏘️ Filtrar por Grupo",
    ["Todos"] + grupos_disponibles,
    index=0
)

if grupo_sel != "Todos":
    df_filtrado_grupo = df_filtrado_mes[df_filtrado_mes["Grupo"] == grupo_sel]
else:
    df_filtrado_grupo = df_filtrado_mes

# --- Publicador ---
publicadores_disponibles = sorted(df_filtrado_grupo["Publicador"].unique().tolist())
publicadores_sel = st.sidebar.multiselect(
    "👤 Filtrar por Publicador",
    publicadores_disponibles,
    default=[],
    placeholder="Todos los publicadores"
)

if publicadores_sel:
    df_filtrado = df_filtrado_grupo[df_filtrado_grupo["Publicador"].isin(publicadores_sel)]
else:
    df_filtrado = df_filtrado_grupo

# --- ORDEN PERSONALIZADO SEGÚN LA FOTO DEL USUARIO ---
orden_publicadores = [
    "Delis Requena",
    "Esteban Requena",
    "David Requena",
    "Erosilda Arbelaez",
    "Anderson Melendez (Aux.)",
    "Lina Sanchez",
    "Teonilda Porto",
    "Diana Maestre",
    "Yenfer Ojeda (S. de grupo)",
    "Yolmarys Ojeda",
    "Mario Arbelaez",
    "Zenaida Jimenez",
    "Ender Requena",
    "Jose Arbelaez",
    "Elizandrit Rebolledo"
]

if not df_filtrado.empty:
    def get_sort_index(val):
        try:
            return orden_publicadores.index(val)
        except ValueError:
            return len(orden_publicadores)
    df_filtrado = df_filtrado.copy()
    df_filtrado["sort_idx"] = df_filtrado["Publicador"].apply(get_sort_index)
    df_filtrado = df_filtrado.sort_values(by=["sort_idx", "Publicador"]).drop(columns=["sort_idx"])

# Etiqueta dinámica
if meses_seleccionados:
    etiqueta_periodo = meses_seleccionados[0] if len(meses_seleccionados) == 1 else f"{len(meses_seleccionados)} Meses"
else:
    etiqueta_periodo = ""

filtro_label = etiqueta_periodo
if grupo_sel != "Todos":
    filtro_label += f" · {grupo_sel}"
if publicadores_sel:
    filtro_label += f" · {len(publicadores_sel)} pub."

# ========================= CÁLCULO DE MÉTRICAS =========================
df_activos_mes = df_filtrado[df_filtrado["Privilegios"] != "Inactivo"]
df_activos_año = df_filtrado_año[df_filtrado_año["Privilegios"] != "Inactivo"]

total_informes_mes = len(df_activos_mes)
total_horas_mes = df_activos_mes["Horas"].sum()
total_cursos_mes = df_activos_mes["Cursos Bíblicos"].sum()

# Nuevas métricas
num_meses = len(meses_seleccionados) if meses_seleccionados else 1
promedio_horas = total_horas_mes / total_informes_mes if total_informes_mes > 0 else 0

total_informes_año = len(df_activos_año)
total_horas_año = df_activos_año["Horas"].sum()

# --- DELTAS ---
delta_inf = ""
delta_hrs = ""
delta_cur = ""

if len(meses_seleccionados) == 1:
    mes_actual_sel = meses_seleccionados[0]
    try:
        y, m = map(int, mes_actual_sel.split('-'))
        mes_ant_str = f"{y-1}-12" if m == 1 else f"{y}-{m-1:02d}"
    except:
        mes_ant_str = None
    
    if mes_ant_str:
        df_mes_ant = df[df["Mes"] == mes_ant_str]
        if grupo_sel != "Todos":
            df_mes_ant = df_mes_ant[df_mes_ant["Grupo"] == grupo_sel]
        if publicadores_sel:
            df_mes_ant = df_mes_ant[df_mes_ant["Publicador"].isin(publicadores_sel)]
            
        df_activos_mes_ant = df_mes_ant[df_mes_ant["Privilegios"] != "Inactivo"]
        total_inf_ant = len(df_activos_mes_ant)
        total_hrs_ant = df_activos_mes_ant["Horas"].sum()
        total_cur_ant = df_activos_mes_ant["Cursos Bíblicos"].sum()
        
        def format_delta(actual, anterior):
            if anterior == 0: return ('▲ 100%', 'delta-up') if actual > 0 else ('', '')
            cambio = ((actual - anterior) / anterior) * 100
            if cambio > 0: return (f'▲ {cambio:.0f}%', 'delta-up')
            if cambio < 0: return (f'▼ {abs(cambio):.0f}%', 'delta-down')
            return ('=', '')

        d_inf_v, d_inf_c = format_delta(total_informes_mes, total_inf_ant)
        d_hrs_v, d_hrs_c = format_delta(total_horas_mes, total_hrs_ant)
        d_cur_v, d_cur_c = format_delta(total_cursos_mes, total_cur_ant)
        
        if d_inf_v: delta_inf = f'<div class="metric-delta {d_inf_c}">{d_inf_v}</div>'
        if d_hrs_v: delta_hrs = f'<div class="metric-delta {d_hrs_c}">{d_hrs_v}</div>'
        if d_cur_v: delta_cur = f'<div class="metric-delta {d_cur_c}">{d_cur_v}</div>'

# Renderizado de Tarjetas
def render_card(label, value, icon, icon_class, footer, delta=""):
    return f"""<div class="metric-card">
<div class="metric-icon-circle {icon_class}">{icon}</div>
<div class="metric-label">{label}</div>
<div class="metric-value-container">
<div class="metric-value">{value}</div>
{delta}
</div>
<div class="metric-footer">{footer}</div>
</div>"""

cards_html = f"""<div class="dashboard-grid">
{render_card("Informes", total_informes_mes, "📄", "icon-teal", etiqueta_periodo, delta_inf)}
{render_card("Horas", total_horas_mes, "🕒", "icon-blue", etiqueta_periodo, delta_hrs)}
{render_card("Cursos", total_cursos_mes, "📖", "icon-orange", etiqueta_periodo, delta_cur)}
{render_card("Horas Año", total_horas_año, "🚀", "icon-purple", año_servicio_sel if año_servicio_sel != 'Todos' else 'Histórico')}
{render_card("Informes Año", total_informes_año, "📅", "icon-rose", año_servicio_sel if año_servicio_sel != 'Todos' else 'Histórico')}
</div>"""

st.markdown(cards_html, unsafe_allow_html=True)

# ========================= TABLA =========================
st.divider()
st.markdown('<div class="section-title">📊 Datos del Periodo</div>', unsafe_allow_html=True)

# Columnas visibles (sin ID ni publisher_id internos)
columnas_visibles = ["Publicador", "Grupo", "Mes", "Privilegios", "Cursos Bíblicos", "Informe", "Notas"]
st.dataframe(
    df_filtrado[columnas_visibles],
    use_container_width=True,
    hide_index=True,
    height=min(520, max(260, 44 * (len(df_filtrado) + 1))),
)



# ========================= GRÁFICO 1: Horas por Publicador =========================
st.divider()
st.markdown('<div class="section-title">🕒 Horas del Mes por Publicador</div>', unsafe_allow_html=True)

df_graficos = df_filtrado[df_filtrado["Privilegios"] != "Inactivo"].copy()

if not df_graficos.empty:
    df_horas_pub = df_graficos.groupby("Publicador", as_index=False)["Horas"].sum()
    df_horas_pub = df_horas_pub.sort_values("Horas", ascending=False)
    # Truncar nombres para que no se superpongan en celulares
    df_horas_pub["Publicador_Corto"] = df_horas_pub["Publicador"].apply(lambda x: nombre_corto(x, 16))

    # Gradient colors from teal to emerald based on value
    n_bars = len(df_horas_pub)
    bar_colors = []
    for i in range(n_bars):
        ratio = i / max(n_bars - 1, 1)
        r = int(0 + ratio * 16)
        g = int(200 - ratio * 55)
        b = int(150 - ratio * 21)
        bar_colors.append(f"rgb({r},{g},{b})")

    fig1 = go.Figure()
    fig1.add_trace(go.Bar(
        x=df_horas_pub["Publicador_Corto"],
        y=df_horas_pub["Horas"],
        text=df_horas_pub["Horas"].apply(lambda x: f"<b>{x}</b>"),
        textposition="outside",
        textfont=dict(size=14, color="#00e8a2", family="Outfit"),
        customdata=df_horas_pub["Publicador"],
        marker=dict(
            color=bar_colors,
            line=dict(width=0),
            opacity=0.92
        ),
        hovertemplate="<b>%{customdata}</b><br>Horas: %{y}<extra></extra>"
    ))
    max_h_pub = df_horas_pub["Horas"].max() if not df_horas_pub.empty else 10
    fig1.update_layout(
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Outfit", size=13, color="#ffffff"),
        showlegend=False,
        xaxis=dict(
            showgrid=False,
            title="",
            tickfont=dict(size=11, color="#c8d6e5", family="Outfit"),
            tickangle=-25 if n_bars > 6 else 0,
            categoryorder="total descending"
        ),
        yaxis=dict(
            showgrid=True,
            gridcolor="rgba(255,255,255,0.04)",
            gridwidth=1,
            title="",
            range=[0, max_h_pub * 1.22],
            tickfont=dict(size=11, color="#8899aa", family="Outfit"),
            zeroline=False
        ),
        margin=dict(t=25, b=60, l=45, r=25),
        height=380,
        bargap=0.25,
        transition=dict(duration=800, easing="cubic-in-out")
    )
    aplicar_layout_premium(fig1, height=390, margin=dict(t=28, b=70, l=48, r=28), mobile=n_bars > 7)
    st.plotly_chart(fig1, use_container_width=True, config=PLOT_CONFIG)
else:
    st.info("No hay datos de publicadores activos en este periodo.")

# ========================= GRÁFICO 2: Evolución de Horas Totales por Mes =========================
st.divider()
st.markdown('<div class="section-title">📈 Evolución de Horas del Grupo por Mes</div>', unsafe_allow_html=True)

df_evolucion_base = df.copy()
if grupo_sel != "Todos":
    df_evolucion_base = df_evolucion_base[df_evolucion_base["Grupo"] == grupo_sel]
if publicadores_sel:
    df_evolucion_base = df_evolucion_base[df_evolucion_base["Publicador"].isin(publicadores_sel)]

df_evolucion = df_evolucion_base[df_evolucion_base["Privilegios"] != "Inactivo"].copy()
if not df_evolucion.empty:
    df_mensual = df_evolucion.groupby("Mes", as_index=False)["Horas"].sum()
    df_mensual = df_mensual.sort_values("Mes")

    fig2 = go.Figure()
    # Gradient area fill behind the line
    fig2.add_trace(go.Scatter(
        x=df_mensual["Mes"],
        y=df_mensual["Horas"],
        mode='lines',
        line=dict(color="rgba(59,130,246,0)", width=0),
        fill='tozeroy',
        fillcolor='rgba(59,130,246,0.07)',
        showlegend=False,
        hoverinfo='skip'
    ))
    # Main line with glow effect
    fig2.add_trace(go.Scatter(
        x=df_mensual["Mes"],
        y=df_mensual["Horas"],
        mode='lines',
        line=dict(color="rgba(96,165,250,0.3)", width=12, shape='spline'),
        showlegend=False,
        hoverinfo='skip'
    ))
    fig2.add_trace(go.Scatter(
        x=df_mensual["Mes"],
        y=df_mensual["Horas"],
        mode='lines+markers+text',
        text=df_mensual["Horas"].apply(lambda x: f"<b>{x}</b>"),
        textposition="top center",
        textfont=dict(size=16, color="#93c5fd", family="Outfit"),
        line=dict(color="#3b82f6", width=3.5, shape='spline'),
        marker=dict(
            size=12,
            color="#3b82f6",
            line=dict(width=3, color="#0f111a"),
            symbol="circle"
        ),
        name="Horas",
        hovertemplate="<b>%{x}</b><br>Horas: %{y}<extra></extra>"
    ))
    max_h_ev = df_mensual["Horas"].max() if not df_mensual.empty else 10
    fig2.update_layout(
         plot_bgcolor="rgba(0,0,0,0)",
         paper_bgcolor="rgba(0,0,0,0)",
         font=dict(family="Outfit", size=13, color="#ffffff"),
         xaxis=dict(
             showgrid=False,
             title="",
             type="category",
             range=[-0.5, len(df_mensual) - 0.5],
             tickfont=dict(size=12, color="#c8d6e5", family="Outfit"),
             tickangle=0
         ),
         yaxis=dict(
             showgrid=True,
             gridcolor="rgba(255,255,255,0.04)",
             gridwidth=1,
             title="",
             range=[0, max_h_ev * 1.25],
             tickfont=dict(size=11, color="#8899aa", family="Outfit"),
             zeroline=False
         ),
         margin=dict(t=30, b=40, l=45, r=25),
         height=340,
         showlegend=False,
         transition=dict(duration=900, easing="cubic-in-out")
     )
    aplicar_layout_premium(fig2, height=350, margin=dict(t=30, b=46, l=48, r=28))
    st.plotly_chart(fig2, use_container_width=True, config=PLOT_CONFIG)
else:
    st.info("No hay suficientes datos para mostrar la evolución.")

# ========================= GRÁFICO 3: Desempeño de Precursores =========================
st.divider()

col_t1, col_t2 = st.columns([1.6, 1.2])
with col_t1:
    st.markdown('<div class="section-title">🏆 Desempeño de Precursores</div>', unsafe_allow_html=True)
with col_t2:
    st.markdown('<div style="height: 35px;"></div>', unsafe_allow_html=True)
    tipo_prec = st.selectbox(
        "Tipo de precursor",
        options=["Todos", "Precursores Regulares", "Precursores Auxiliares"],
        index=0,
        label_visibility="collapsed"
    )

df_precursores = df_filtrado[df_filtrado["Privilegios"].str.contains("precursor", case=False, na=False)].copy()

if tipo_prec == "Precursores Regulares":
    df_precursores = df_precursores[df_precursores["Privilegios"].str.contains("regular", case=False, na=False)]
elif tipo_prec == "Precursores Auxiliares":
    df_precursores = df_precursores[df_precursores["Privilegios"].str.contains("auxiliar", case=False, na=False)]

if not df_precursores.empty:
    df_prec_resumen = df_precursores.groupby("Publicador", as_index=False).agg({
        "Horas": "sum",
        "Cursos Bíblicos": "sum"
    })
    df_prec_resumen = df_prec_resumen.sort_values("Horas", ascending=True)
    # Truncar nombres para que no se corten o superpongan en celulares
    df_prec_resumen["Publicador_Corto"] = df_prec_resumen["Publicador"].apply(lambda x: nombre_corto(x, 18))

    # Gradient purple colors for each bar
    n_prec = len(df_prec_resumen)
    prec_colors = []
    for i in range(n_prec):
        ratio = i / max(n_prec - 1, 1)
        r = int(139 - ratio * 48)
        g = int(92 - ratio * 39)
        b = int(246 - ratio * 64)
        prec_colors.append(f"rgb({r},{g},{b})")

    fig3 = go.Figure()
    fig3.add_trace(go.Bar(
        y=df_prec_resumen["Publicador_Corto"],
        x=df_prec_resumen["Horas"],
        orientation='h',
        marker=dict(
            color=prec_colors,
            line=dict(width=0),
            opacity=0.9
        ),
        text=df_prec_resumen["Horas"].apply(lambda x: f"<b>{x}h</b>"),
        textposition="outside",
        textfont=dict(size=13, color="#c4b5fd", family="Outfit"),
        customdata=df_prec_resumen["Publicador"],
        name="Horas",
        hovertemplate="<b>%{customdata}</b><br>Horas: %{x}<extra></extra>"
    ))
    max_h_prec = df_prec_resumen["Horas"].max() if not df_prec_resumen.empty else 10
    fig3.update_layout(
        height=max(240, n_prec * 48 + 80),
        title=dict(
            text="<b>Horas de Precursores</b>",
            font=dict(color="#ffffff", size=16, family="Outfit"),
            x=0.5,
            xanchor="center"
        ),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Outfit", color="white", size=13),
        xaxis=dict(
            title="",
            tickfont=dict(color="#8899aa", size=11, family="Outfit"),
            gridcolor="rgba(255,255,255,0.04)",
            zerolinecolor="rgba(255,255,255,0.04)",
            range=[0, max_h_prec * 1.22]
        ),
        yaxis=dict(
            tickfont=dict(color="#c8d6e5", size=11, family="Outfit"),
            gridcolor="rgba(255,255,255,0.04)",
            zerolinecolor="rgba(255,255,255,0.04)"
        ),
        showlegend=False,
        margin=dict(t=45, b=25, l=85, r=45),
        bargap=0.2,
        transition=dict(duration=800, easing="cubic-in-out")
    )
    aplicar_layout_premium(fig3, height=max(260, n_prec * 50 + 90), margin=dict(t=48, b=30, l=95, r=54), mobile=n_prec > 7)
    st.plotly_chart(fig3, use_container_width=True, config=PLOT_CONFIG)

    col_p1, col_p2 = st.columns(2)
    with col_p1:
        total_horas_prec = df_prec_resumen["Horas"].sum()
        st.metric("Total Horas Precursores", f"{total_horas_prec} hrs")
    with col_p2:
        total_cursos_prec = df_prec_resumen["Cursos Bíblicos"].sum()
        st.metric("Total Cursos Bíblicos", total_cursos_prec)
else:
    st.info("No hay datos de precursores en este periodo. Registra informes con privilegio 'Precursor Regular' o 'Precursor Auxiliar' para ver datos aquí.")

# ========================= GRÁFICO 4: Gráfico Circular =========================
st.divider()
st.markdown('<div class="section-title">🍩 Distribución de Horas por Publicador</div>', unsafe_allow_html=True)

df_circular = df_filtrado[df_filtrado["Horas"] > 0].copy()
if not df_circular.empty:
    df_circular = df_circular.groupby("Publicador", as_index=False)["Horas"].sum().sort_values("Horas", ascending=False)
    # Truncar nombres para que no se superpongan o salgan del gráfico
    df_circular["Publicador_Corto"] = df_circular["Publicador"].apply(lambda x: nombre_corto(x, 16))
    # Premium color palette
    premium_colors = [
        "#00c896", "#3b82f6", "#8b5cf6", "#f59e0b", "#f43f5e",
        "#06b6d4", "#10b981", "#6366f1", "#ec4899", "#14b8a6",
        "#a855f7", "#22d3ee", "#84cc16", "#e879f9", "#fb923c"
    ]
    fig_pie = go.Figure()
    fig_pie.add_trace(go.Pie(
        labels=df_circular["Publicador_Corto"],
        values=df_circular["Horas"],
        hole=0.62,
        marker=dict(
            colors=PREMIUM_COLORS[:len(df_circular)],
            line=dict(color='#0f111a', width=2.5)
        ),
        customdata=df_circular["Publicador"],
        textposition='outside',
        textinfo='percent+label',
        textfont=dict(size=11, color="#c8d6e5", family="Outfit"),
        hovertemplate="<b>%{customdata}</b><br>Horas: %{value}<br>%{percent}<extra></extra>",
        pull=[0.03] * len(df_circular),
        rotation=90
    ))
    # Add center annotation
    total_hrs_pie = df_circular["Horas"].sum()
    fig_pie.update_layout(
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Outfit", size=12, color="#ffffff"),
        showlegend=False,
        margin=dict(t=35, b=35, l=40, r=40),
        height=380,
        annotations=[
            dict(
                text=f"<b>{total_hrs_pie}</b><br><span style='font-size:11px;color:#8899aa'>Horas<br>Totales</span>",
                x=0.5, y=0.5,
                font=dict(size=22, color="#00c896", family="Outfit"),
                showarrow=False
            )
        ],
        transition=dict(duration=700, easing="cubic-in-out")
    )
    aplicar_layout_premium(fig_pie, height=390, margin=dict(t=34, b=34, l=36, r=36), mobile=len(df_circular) > 7)
    st.plotly_chart(fig_pie, use_container_width=True, config=PLOT_CONFIG)
else:
    st.info("No hay suficientes datos de horas para generar el gráfico circular.")

# ========================= EXPORTAR =========================
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io

st.divider()
st.markdown('<div class="section-title">📥 Exportar Reporte</div>', unsafe_allow_html=True)

buffer = io.BytesIO()

# Preparar y limpiar los datos exactos del reporte
df_export = df_filtrado[["Publicador", "Privilegios", "Mes", "Informe", "Cursos Bíblicos"]].copy()
df_export.columns = ["Nombre y apellido", "Privilegio (publicador, precursor, etc.)", "Mes", "Informe", "Cursos biblicos"]

# Lógica para Inactivos: dejar celdas en blanco en vez de valores por defecto
inactivos_mask = df_export["Privilegio (publicador, precursor, etc.)"].str.lower() == "inactivo"
df_export["Mes"] = df_export["Mes"].astype(object)
df_export["Informe"] = df_export["Informe"].astype(object)
df_export["Cursos biblicos"] = df_export["Cursos biblicos"].astype(object)
df_export.loc[inactivos_mask, "Mes"] = ""
df_export.loc[inactivos_mask, "Informe"] = ""
df_export.loc[inactivos_mask, "Cursos biblicos"] = ""

# Lógica para Cursos Bíblicos: si es 0, dejarlo vacío (como en la imagen de Excel original)
def clean_cursos(x):
    if pd.isna(x) or str(x).strip() == "": return ""
    try:
        val = int(float(x))
        return "" if val == 0 else val
    except:
        return x

df_export["Cursos biblicos"] = df_export["Cursos biblicos"].apply(clean_cursos)
df_export.loc[inactivos_mask, "Cursos biblicos"] = ""

# Guardar en buffer usando el motor de openpyxl
with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
    # Escribir el DataFrame desde A2 (startrow=1, startcol=0)
    df_export.to_excel(writer, index=False, sheet_name='Informes', startrow=1, startcol=0)

    workbook = writer.book
    worksheet = writer.sheets['Informes']

    # --- Estilos reutilizables ---
    borde_grueso_ext = Border(
        left=Side(style='medium', color='1E3A5F'),
        right=Side(style='medium', color='1E3A5F'),
        top=Side(style='medium', color='1E3A5F'),
        bottom=Side(style='medium', color='1E3A5F')
    )
    borde_fino = Border(
        left=Side(style='thin', color='B0C4D8'),
        right=Side(style='thin', color='B0C4D8'),
        top=Side(style='thin', color='B0C4D8'),
        bottom=Side(style='thin', color='B0C4D8')
    )
    fill_header  = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    fill_par     = PatternFill(start_color='EBF3FF', end_color='EBF3FF', fill_type='solid')
    fill_impar   = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    font_header  = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_data    = Font(name='Calibri', size=10, color='1A1A2E')

    # --- Configurar todas las columnas con el mismo tamaño (Ancho: 22) ---
    num_cols = len(df_export.columns)
    for col_idx in range(1, num_cols + 1):
        col_letter = worksheet.cell(row=2, column=col_idx).column_letter
        worksheet.column_dimensions[col_letter].width = 22

    # --- Título en fila 1 ---
    titulo_texto = f"Grupo {grupo_sel}" if grupo_sel != "Todos" else f"Reporte General — {filtro_label}"
    last_col_letter = worksheet.cell(row=2, column=num_cols).column_letter
    worksheet.merge_cells(f'A1:{last_col_letter}1')
    titulo_cell = worksheet['A1']
    titulo_cell.value = titulo_texto
    titulo_cell.fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    titulo_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    worksheet.row_dimensions[1].height = 32

    # --- Encabezados (fila 2) y datos (fila 3 en adelante) ---
    num_filas = len(df_export)
    header_row = 2
    data_start = 3

    for r_idx in range(header_row, data_start + num_filas):
        worksheet.row_dimensions[r_idx].height = 20 if r_idx >= data_start else 24
        for c_idx in range(1, num_cols + 1):
            cell = worksheet.cell(row=r_idx, column=c_idx)
            if r_idx == header_row:
                cell.font = font_header
                cell.fill = fill_header
                cell.border = borde_grueso_ext
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                es_par = (r_idx - data_start) % 2 == 0
                cell.fill = fill_par if es_par else fill_impar
                cell.font = font_data
                cell.border = borde_fino
                if c_idx == 4: # Informe (columna D)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Filtros automáticos sobre la fila de encabezados ---
    last_data_letter = worksheet.cell(row=header_row, column=num_cols).column_letter
    worksheet.auto_filter.ref = f'A{header_row}:{last_data_letter}{data_start + num_filas - 1}'

    # --- Congelar la fila de encabezados ---
    worksheet.freeze_panes = f'A{data_start}'

    # --- Cuadrícula nativa visible ---
    worksheet.sheet_view.showGridLines = True

buffer.seek(0)
excel_filename = f"Reporte_{(etiqueta_periodo or 'general').replace(' ', '_')}.xlsx"
col_b1 = st.container()
with col_b1:
    st.download_button(
        label="📥 Descargar Excel",
        data=buffer.getvalue(),
        file_name=excel_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

# ========================= ANÁLISIS IA =========================
st.divider()
st.markdown('<div class="section-title">🤖 Análisis Inteligente (IA)</div>', unsafe_allow_html=True)

with st.container():
    col_ia_1, col_ia_2 = st.columns([1, 2])
    with col_ia_1:
        ai_opcion = st.radio("Motor de IA", ["Google Gemini", "OpenAI"], horizontal=True)
        modelo_visible = OPENAI_MODEL if ai_opcion == "OpenAI" else GEMINI_MODEL
        st.caption(f"Modelo activo: {modelo_visible}")
        btn_ia = st.button("✨ Generar Reporte con IA", use_container_width=True, type="primary")

if btn_ia:
    with st.status("🧠 Analizando datos y generando reporte...", expanded=True) as status:
        df_activos = df_filtrado[df_filtrado["Privilegios"] != "Inactivo"]
        inactivos_nombres = df_filtrado[df_filtrado["Privilegios"] == "Inactivo"]["Publicador"].tolist()
        
        mask_no_predico = df_activos["Informe"].str.contains("no predic", case=False, na=False) | (df_activos["Horas"] == 0)
        sin_informe = df_activos[mask_no_predico]["Publicador"].tolist()
        
        total_pubs = len(df_activos)
        total_horas_ia = df_activos["Horas"].sum()
        total_cursos_ia = df_activos["Cursos Bíblicos"].sum()
        
        # Comparación con mes anterior
        total_inf_ant = 0
        total_hrs_ant = 0
        total_cur_ant = 0
        comparacion_disponible = False
        mes_ant_str = None
        
        if meses_seleccionados and len(meses_seleccionados) == 1:
            mes_actual_sel = meses_seleccionados[0]
            try:
                y, m = map(int, mes_actual_sel.split('-'))
                mes_ant_str = f"{y-1}-12" if m == 1 else f"{y}-{m-1:02d}"
            except:
                mes_ant_str = None
            
            if mes_ant_str:
                df_mes_ant = df[df["Mes"] == mes_ant_str]
                if grupo_sel != "Todos":
                    df_mes_ant = df_mes_ant[df_mes_ant["Grupo"] == grupo_sel]
                if publicadores_sel:
                    df_mes_ant = df_mes_ant[df_mes_ant["Publicador"].isin(publicadores_sel)]
                    
                df_activos_mes_ant = df_mes_ant[df_mes_ant["Privilegios"] != "Inactivo"]
                total_inf_ant = len(df_activos_mes_ant)
                total_hrs_ant = df_activos_mes_ant["Horas"].sum()
                total_cur_ant = df_activos_mes_ant["Cursos Bíblicos"].sum()
                comparacion_disponible = True

        comparacion_text = ""
        if comparacion_disponible:
            comparacion_text = f"""
            COMPARACIÓN CON EL MES ANTERIOR ({mes_ant_str}):
            - Informes presentados el mes pasado: {total_inf_ant} (Mes actual: {total_pubs})
            - Horas totales el mes pasado: {total_hrs_ant} (Mes actual: {total_horas_ia})
            - Cursos bíblicos el mes pasado: {total_cur_ant} (Mes actual: {total_cursos_ia})
            """

        precursores = df_activos[df_activos["Privilegios"].str.contains("precursor", case=False, na=False)]
        prec_text = "\n".join([f"- {row['Publicador']} ({row['Privilegios']}): {row['Horas']} horas, {row['Cursos Bíblicos']} cursos" for _, row in precursores.iterrows()])
        
        prompt = f"""Eres un analista de datos experto. Analiza este resumen del periodo {año_servicio_sel} ({etiqueta_periodo}).
        
        Datos del grupo actual:
        - Total publicadores activos: {total_pubs}
        - Horas totales: {total_horas_ia}
        - Cursos bíblicos: {total_cursos_ia}
        - Inactivos: {', '.join(inactivos_nombres) if inactivos_nombres else 'Ninguno'}
        - Sin actividad este mes: {', '.join(sin_informe) if sin_informe else 'Ninguno'}
        - Precursores:
        {prec_text if not precursores.empty else 'No hay precursores.'}
        
        {comparacion_text}
        
        Genera un informe estructurado con las siguientes secciones:
        1. RESUMEN EJECUTIVO (Analiza los datos sobresalientes del mes actual)
        2. COMPARACIÓN CON EL MES ANTERIOR (Menciona si aumentaron/disminuyeron las horas, cursos e informes y qué significa esto)
        3. PUNTOS FUERTES Y LOGROS (Destaca lo positivo del grupo)
        4. ÁREAS DE MEJORA Y SUGERENCIAS (Identifica qué aspectos se pueden mejorar y da consejos prácticos)
        
        Sé muy profesional, motivador y preciso. Evita respuestas genéricas. Utiliza Markdown limpio."""

        try:
            resultado = ""
            if ai_opcion == "OpenAI":
                api_key = os.getenv("OPENAI_API_KEY")
                if api_key:
                    client = OpenAI(api_key=api_key)
                    res = client.chat.completions.create(
                        model=OPENAI_MODEL,
                        messages=[{"role": "user", "content": prompt}],
                        temperature=0.35,
                    )
                    resultado = res.choices[0].message.content
                else:
                    st.error("Falta API KEY de OpenAI.")
            else:
                api_key = os.getenv("GEMINI_API_KEY")
                if api_key:
                    genai.configure(api_key=api_key)
                    model = genai.GenerativeModel(GEMINI_MODEL)
                    res = model.generate_content(prompt)
                    resultado = res.text
                else:
                    st.error("Falta API KEY de Gemini.")
            
            if resultado:
                modelo_usado = OPENAI_MODEL if ai_opcion == "OpenAI" else GEMINI_MODEL
                # Proceder a generar PDF con el informe
                from fpdf import FPDF
                import tempfile
                import uuid
                
                # Obtener ruta de la carpeta "analisis de informe"
                app_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                workspace_dir = os.path.dirname(app_dir)
                pdf_dir = os.path.join(workspace_dir, "analisis de informe")
                os.makedirs(pdf_dir, exist_ok=True)
                
                mes_sel = meses_seleccionados[0] if meses_seleccionados else datetime.datetime.now().strftime("%Y-%m")
                pdf_filename = f"informe mes {mes_sel}.pdf"
                pdf_path = os.path.join(pdf_dir, pdf_filename)
                
                # ============ GENERAR GRÁFICOS PREMIUM PARA PDF ============
                temp_dir = tempfile.gettempdir()
                img_paths = {}
                
                # Colores base para PDF (fondo oscuro premium)
                PDF_BG = "#0f172a"
                PDF_PLOT_BG = "rgba(15,23,42,0.95)"
                PDF_GRID = "rgba(148,163,184,0.12)"
                PDF_TEXT = "#e2e8f0"
                PDF_MUTED = "#94a3b8"
                
                def _save_premium_fig(fig_obj, name, w=900, h=420):
                    """Guarda un gráfico Plotly premium como imagen PNG de alta resolución."""
                    try:
                        path = os.path.join(temp_dir, f"{name}_{uuid.uuid4().hex}.png")
                        fig_obj.write_image(path, width=w, height=h, scale=3)
                        return path
                    except Exception as e:
                        return None
                
                # --- FIG 1 PDF: Barras de Horas por Publicador ---
                try:
                    df_g1 = df_graficos.groupby("Publicador", as_index=False)["Horas"].sum().sort_values("Horas", ascending=False) if not df_graficos.empty else pd.DataFrame()
                    if not df_g1.empty:
                        n1 = len(df_g1)
                        colors1 = []
                        for i in range(n1):
                            ratio = i / max(n1 - 1, 1)
                            cr = int(0 + ratio * 16)
                            cg = int(220 - ratio * 60)
                            cb = int(160 - ratio * 25)
                            colors1.append(f"rgb({cr},{cg},{cb})")
                        
                        fig1_pdf = go.Figure()
                        fig1_pdf.add_trace(go.Bar(
                            x=df_g1["Publicador"],
                            y=df_g1["Horas"],
                            text=df_g1["Horas"].apply(lambda x: f"<b>{x}</b>"),
                            textposition="outside",
                            textfont=dict(size=18, color="#00e8a2", family="Arial Black"),
                            marker=dict(color=colors1, line=dict(width=0), opacity=0.92),
                        ))
                        max1 = df_g1["Horas"].max()
                        fig1_pdf.update_layout(
                            plot_bgcolor=PDF_PLOT_BG, paper_bgcolor=PDF_BG,
                            font=dict(family="Arial", size=14, color=PDF_TEXT),
                            showlegend=False,
                            title=dict(
                                text="<b>🕒 Horas del Mes por Publicador</b>",
                                font=dict(size=22, color="#ffffff", family="Arial Black"),
                                x=0.5, xanchor="center", y=0.96
                            ),
                            xaxis=dict(
                                showgrid=False, title="",
                                tickfont=dict(size=13, color=PDF_TEXT, family="Arial"),
                                tickangle=-20 if n1 > 6 else 0,
                                categoryorder="total descending"
                            ),
                            yaxis=dict(
                                showgrid=True, gridcolor=PDF_GRID, gridwidth=1,
                                title="", range=[0, max1 * 1.25],
                                tickfont=dict(size=13, color=PDF_MUTED, family="Arial"),
                                zeroline=False
                            ),
                            margin=dict(t=60, b=80, l=55, r=30),
                            bargap=0.25
                        )
                        img_paths['fig1'] = _save_premium_fig(fig1_pdf, "fig1_pdf", w=950, h=440)
                except:
                    pass
                
                # --- FIG 2 PDF: Evolución de Horas por Mes ---
                try:
                    df_ev_pdf = df_evolucion_base[df_evolucion_base["Privilegios"] != "Inactivo"].copy() if 'df_evolucion_base' in dir() else pd.DataFrame()
                    if not df_ev_pdf.empty:
                        df_m_pdf = df_ev_pdf.groupby("Mes", as_index=False)["Horas"].sum().sort_values("Mes")
                        if not df_m_pdf.empty:
                            fig2_pdf = go.Figure()
                            # Area fill
                            fig2_pdf.add_trace(go.Scatter(
                                x=df_m_pdf["Mes"], y=df_m_pdf["Horas"],
                                mode='lines', line=dict(color="rgba(59,130,246,0)", width=0),
                                fill='tozeroy', fillcolor='rgba(59,130,246,0.12)',
                                showlegend=False, hoverinfo='skip'
                            ))
                            # Glow line
                            fig2_pdf.add_trace(go.Scatter(
                                x=df_m_pdf["Mes"], y=df_m_pdf["Horas"],
                                mode='lines', line=dict(color="rgba(96,165,250,0.25)", width=14, shape='spline'),
                                showlegend=False, hoverinfo='skip'
                            ))
                            # Main line + markers
                            fig2_pdf.add_trace(go.Scatter(
                                x=df_m_pdf["Mes"], y=df_m_pdf["Horas"],
                                mode='lines+markers+text',
                                text=df_m_pdf["Horas"].apply(lambda x: f"<b>{x}</b>"),
                                textposition="top center",
                                textfont=dict(size=18, color="#93c5fd", family="Arial Black"),
                                line=dict(color="#3b82f6", width=4, shape='spline'),
                                marker=dict(size=14, color="#3b82f6", line=dict(width=3, color=PDF_BG), symbol="circle"),
                                name="Horas"
                            ))
                            max2 = df_m_pdf["Horas"].max()
                            fig2_pdf.update_layout(
                                plot_bgcolor=PDF_PLOT_BG, paper_bgcolor=PDF_BG,
                                font=dict(family="Arial", size=14, color=PDF_TEXT),
                                title=dict(
                                    text="<b>📈 Evolución de Horas del Grupo por Mes</b>",
                                    font=dict(size=22, color="#ffffff", family="Arial Black"),
                                    x=0.5, xanchor="center", y=0.96
                                ),
                                xaxis=dict(
                                    showgrid=False, title="", type="category",
                                    range=[-0.5, len(df_m_pdf) - 0.5],
                                    tickfont=dict(size=14, color=PDF_TEXT, family="Arial")
                                ),
                                yaxis=dict(
                                    showgrid=True, gridcolor=PDF_GRID, gridwidth=1,
                                    title="", range=[0, max2 * 1.3],
                                    tickfont=dict(size=13, color=PDF_MUTED, family="Arial"),
                                    zeroline=False
                                ),
                                margin=dict(t=60, b=55, l=60, r=30),
                                showlegend=False
                            )
                            img_paths['fig2'] = _save_premium_fig(fig2_pdf, "fig2_pdf", w=950, h=420)
                except:
                    pass
                
                # --- FIG 3 PDF: Desempeño de Precursores ---
                try:
                    df_prec_pdf = df_precursores.groupby("Publicador", as_index=False).agg({"Horas": "sum", "Cursos Bíblicos": "sum"}).sort_values("Horas", ascending=True) if ('df_precursores' in dir() and not df_precursores.empty) else pd.DataFrame()
                    if not df_prec_pdf.empty:
                        n3 = len(df_prec_pdf)
                        colors3 = []
                        for i in range(n3):
                            ratio = i / max(n3 - 1, 1)
                            cr = int(139 - ratio * 48)
                            cg = int(92 - ratio * 39)
                            cb = int(246 - ratio * 64)
                            colors3.append(f"rgb({cr},{cg},{cb})")
                        
                        fig3_pdf = go.Figure()
                        fig3_pdf.add_trace(go.Bar(
                            y=df_prec_pdf["Publicador"], x=df_prec_pdf["Horas"],
                            orientation='h',
                            marker=dict(color=colors3, line=dict(width=0), opacity=0.9),
                            text=df_prec_pdf["Horas"].apply(lambda x: f"<b>{x} hrs</b>"),
                            textposition="outside",
                            textfont=dict(size=16, color="#c4b5fd", family="Arial Black"),
                            name="Horas"
                        ))
                        max3 = df_prec_pdf["Horas"].max()
                        fig3_pdf.update_layout(
                            height=max(300, n3 * 60 + 120),
                            title=dict(
                                text=f"<b>🏆 Desempeño de Precursores</b> — {filtro_label}",
                                font=dict(color="#ffffff", size=22, family="Arial Black"),
                                x=0.5, xanchor="center", y=0.96
                            ),
                            plot_bgcolor=PDF_PLOT_BG, paper_bgcolor=PDF_BG,
                            font=dict(family="Arial", color=PDF_TEXT, size=14),
                            xaxis=dict(
                                title="", tickfont=dict(color=PDF_MUTED, size=13, family="Arial"),
                                gridcolor=PDF_GRID, zerolinecolor=PDF_GRID,
                                range=[0, max3 * 1.25]
                            ),
                            yaxis=dict(
                                tickfont=dict(color=PDF_TEXT, size=14, family="Arial"),
                                gridcolor=PDF_GRID, zerolinecolor=PDF_GRID
                            ),
                            showlegend=False,
                            margin=dict(t=60, b=30, l=140, r=70),
                            bargap=0.3
                        )
                        h3 = max(300, n3 * 60 + 120)
                        img_paths['fig3'] = _save_premium_fig(fig3_pdf, "fig3_pdf", w=950, h=h3)
                except:
                    pass
                
                # --- FIG PIE PDF: Distribución de Horas ---
                try:
                    df_pie_pdf = df_filtrado[df_filtrado["Horas"] > 0].copy()
                    if not df_pie_pdf.empty:
                        premium_cols = [
                            "#00c896", "#3b82f6", "#8b5cf6", "#f59e0b", "#f43f5e",
                            "#06b6d4", "#10b981", "#6366f1", "#ec4899", "#14b8a6",
                            "#a855f7", "#22d3ee", "#84cc16", "#e879f9", "#fb923c"
                        ]
                        total_pie = df_pie_pdf["Horas"].sum()
                        
                        fig_pie_pdf = go.Figure()
                        fig_pie_pdf.add_trace(go.Pie(
                            labels=df_pie_pdf["Publicador"],
                            values=df_pie_pdf["Horas"],
                            hole=0.62,
                            marker=dict(
                                colors=premium_cols[:len(df_pie_pdf)],
                                line=dict(color=PDF_BG, width=3)
                            ),
                            textposition='outside',
                            textinfo='percent+label',
                            textfont=dict(size=14, color=PDF_TEXT, family="Arial"),
                            pull=[0.03] * len(df_pie_pdf),
                            rotation=90
                        ))
                        fig_pie_pdf.update_layout(
                            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor=PDF_BG,
                            font=dict(family="Arial", size=14, color=PDF_TEXT),
                            title=dict(
                                text="<b>🍩 Distribución de Horas por Publicador</b>",
                                font=dict(size=22, color="#ffffff", family="Arial Black"),
                                x=0.5, xanchor="center", y=0.97
                            ),
                            showlegend=False,
                            margin=dict(t=65, b=40, l=90, r=90),
                            annotations=[
                                dict(
                                    text=f"<b>{total_pie}</b><br><span style='font-size:14px;color:#94a3b8'>Horas<br>Totales</span>",
                                    x=0.5, y=0.5,
                                    font=dict(size=32, color="#00c896", family="Arial Black"),
                                    showarrow=False
                                )
                            ]
                        )
                        img_paths['fig_pie'] = _save_premium_fig(fig_pie_pdf, "figpie_pdf", w=900, h=500)
                except:
                    pass
                
                # ============ CONSTRUIR PDF ============
                class PDF(FPDF):
                    def footer(self):
                        self.set_y(-14)
                        self.set_font('Helvetica', 'I', 8)
                        self.set_text_color(120, 130, 150)
                        self.cell(0, 8, self.to_latin1(f"Generado con Inteligencia Artificial · Página {self.page_no()}/{{nb}}"), align='C')
                        
                    def to_latin1(self, text):
                        return text.encode('latin-1', 'replace').decode('latin-1')
                    
                    def draw_section_header(self, text, r=15, g=23, b=42):
                        """Dibuja un encabezado de sección con barra decorativa."""
                        self.set_fill_color(r, g, b)
                        x = self.get_x()
                        y = self.get_y()
                        self.rect(x, y, 190, 8, 'F')
                        self.set_font('Helvetica', 'B', 11)
                        self.set_text_color(255, 255, 255)
                        self.cell(190, 8, self.to_latin1(f"  {text}"), align='L')
                        self.ln(11)
                        self.set_text_color(0, 0, 0)

                def _pdf_footer_clean(self):
                    self.set_y(-14)
                    self.set_font('Helvetica', 'I', 8)
                    self.set_text_color(120, 130, 150)
                    self.cell(0, 8, self.to_latin1(f"Generado con IA - Pagina {self.page_no()}/{{nb}}"), align='C')

                PDF.footer = _pdf_footer_clean

                pdf = PDF()
                pdf.alias_nb_pages()
                pdf.add_page()
                pdf.set_auto_page_break(auto=True, margin=18)
                
                # ---- BANNER PRINCIPAL ----
                # Fondo oscuro premium
                pdf.set_fill_color(15, 23, 42)
                pdf.rect(0, 0, 210, 30, 'F')
                # Línea decorativa verde
                pdf.set_fill_color(0, 200, 150)
                pdf.rect(0, 30, 210, 1.5, 'F')
                
                pdf.set_text_color(255, 255, 255)
                pdf.set_font('Helvetica', 'B', 15)
                pdf.set_y(6)
                pdf.cell(0, 7, pdf.to_latin1("INFORME DE ANÁLISIS MENSUAL CON IA"), align='C')
                pdf.ln(8)
                pdf.set_font('Helvetica', '', 11)
                pdf.set_text_color(200, 214, 229)
                grupo_lbl = f"Grupo: {grupo_sel}" if grupo_sel != "Todos" else "Reporte General"
                pdf.cell(0, 5, pdf.to_latin1(f"{grupo_lbl}  ·  Período: {mes_sel}"), align='C')
                
                pdf.set_fill_color(15, 23, 42)
                pdf.rect(0, 4, 210, 26, 'F')
                pdf.set_text_color(255, 255, 255)
                pdf.set_font('Helvetica', 'B', 15)
                pdf.set_y(6)
                pdf.cell(0, 7, pdf.to_latin1("INFORME DE ANALISIS MENSUAL CON IA"), align='C')
                pdf.ln(8)
                pdf.set_font('Helvetica', '', 10)
                pdf.set_text_color(200, 214, 229)
                pdf.cell(0, 5, pdf.to_latin1(f"{grupo_lbl} - Periodo: {mes_sel} - Modelo IA: {modelo_usado}"), align='C')

                pdf.set_y(38)
                
                # ---- 1. TABLA COMPARATIVA ----
                pdf.draw_section_header("1. TABLA COMPARATIVA DE MÉTRICAS CLAVE")
                
                # Encabezados de tabla con estilo
                pdf.set_fill_color(30, 41, 59)
                pdf.set_text_color(255, 255, 255)
                pdf.set_font('Helvetica', 'B', 9)
                col_widths = [62, 38, 38, 42]
                headers = [" Métrica", "Mes Anterior", "Mes Actual", "Variación"]
                for i, (h, w) in enumerate(zip(headers, col_widths)):
                    pdf.cell(w, 8, pdf.to_latin1(h), border=0, fill=True, align='C' if i > 0 else 'L')
                pdf.ln()
                
                def calc_pct(act, ant):
                    if ant == 0:
                        return "+100%" if act > 0 else "0%"
                    diff = ((act - ant) / ant) * 100
                    return f"{'+' if diff > 0 else ''}{diff:.1f}%"
                    
                metrics = [
                    ("Publicadores Activos", total_inf_ant, total_pubs),
                    ("Horas de Predicación", total_hrs_ant, total_horas_ia),
                    ("Cursos Bíblicos", total_cur_ant, total_cursos_ia)
                ]
                
                pdf.set_font('Helvetica', '', 9)
                for idx, (name, ant, act) in enumerate(metrics):
                    if idx % 2 == 0:
                        pdf.set_fill_color(241, 245, 249)
                    else:
                        pdf.set_fill_color(255, 255, 255)
                    pdf.set_text_color(30, 41, 59)
                    pdf.cell(col_widths[0], 7, pdf.to_latin1(f"  {name}"), border=0, fill=True)
                    pdf.set_text_color(71, 85, 105)
                    pdf.cell(col_widths[1], 7, pdf.to_latin1(str(ant) if comparacion_disponible else "N/A"), border=0, fill=True, align='C')
                    pdf.set_text_color(15, 23, 42)
                    pdf.set_font('Helvetica', 'B', 9)
                    pdf.cell(col_widths[2], 7, pdf.to_latin1(str(act)), border=0, fill=True, align='C')
                    pdf.set_font('Helvetica', '', 9)
                    # Color de variación
                    pct_str = calc_pct(act, ant) if comparacion_disponible else "N/A"
                    if pct_str.startswith('+'):
                        pdf.set_text_color(16, 185, 129)
                    elif pct_str.startswith('-'):
                        pdf.set_text_color(239, 68, 68)
                    else:
                        pdf.set_text_color(71, 85, 105)
                    pdf.set_font('Helvetica', 'B', 9)
                    pdf.cell(col_widths[3], 7, pdf.to_latin1(pct_str), border=0, fill=True, align='C')
                    pdf.set_font('Helvetica', '', 9)
                    pdf.ln()
                
                # Línea separadora
                pdf.set_draw_color(226, 232, 240)
                pdf.line(10, pdf.get_y(), 200, pdf.get_y())
                pdf.ln(6)
                
                # ---- 2. DETALLE DE IA ----
                pdf.draw_section_header("2. ANÁLISIS DETALLADO DE LA IA")
                
                pdf.set_font('Helvetica', '', 9)
                pdf.set_text_color(30, 41, 59)
                
                for line in resultado.split('\n'):
                    line = line.strip()
                    if not line:
                        pdf.ln(1.5)
                        continue
                    
                    if line.startswith('###') or (line.startswith('**') and line.endswith('**')) or line.upper().startswith(('1.', '2.', '3.', '4.')):
                        pdf.ln(2.5)
                        pdf.set_font('Helvetica', 'B', 10)
                        pdf.set_text_color(15, 23, 42)
                        clean_line = line.replace('###', '').replace('**', '').strip()
                        pdf.multi_cell(0, 5, pdf.to_latin1(clean_line), new_x="LMARGIN", new_y="NEXT")
                        pdf.set_font('Helvetica', '', 9)
                        pdf.set_text_color(30, 41, 59)
                    elif line.startswith('-') or line.startswith('*'):
                        clean_line = line.lstrip('-*').strip()
                        pdf.multi_cell(0, 4.5, pdf.to_latin1(f"  · {clean_line}"), new_x="LMARGIN", new_y="NEXT")
                    else:
                        pdf.multi_cell(0, 4.5, pdf.to_latin1(line), new_x="LMARGIN", new_y="NEXT")
                        
                pdf.ln(4)
                
                # ---- 3. GRÁFICOS PREMIUM ----
                # Cada gráfico en su propia página para máxima calidad
                chart_info = [
                    ("fig1",    "Horas del Mes por Publicador"),
                    ("fig2",    "Evolución de Horas del Grupo por Mes"),
                    ("fig3",    "Desempeño de Precursores"),
                    ("fig_pie", "Distribución de Horas por Publicador")
                ]
                
                charts_added = False
                for key, title_lbl in chart_info:
                    path = img_paths.get(key)
                    if path and os.path.exists(path):
                        # Sección de gráficos: título solo la primera vez
                        if not charts_added:
                            pdf.add_page()
                            pdf.draw_section_header("3. ANEXO: GRÁFICOS DE RENDIMIENTO")
                            charts_added = True
                        else:
                            # Verificar espacio: si queda menos de 120mm, nueva página
                            if pdf.get_y() > 145:
                                pdf.add_page()
                        
                        # Subtítulo del gráfico
                        pdf.set_font('Helvetica', 'B', 10)
                        pdf.set_text_color(15, 23, 42)
                        pdf.cell(0, 6, pdf.to_latin1(title_lbl), align='C')
                        pdf.ln(3)
                        
                        # Insertar imagen centrada con borde redondeado visual
                        img_w = 170  # Ancho generoso
                        img_x = (210 - img_w) / 2
                        
                        # Fondo oscuro detrás de la imagen (simula marco)
                        y_before = pdf.get_y()
                        pdf.set_fill_color(15, 23, 42)
                        pdf.set_draw_color(30, 41, 59)
                        
                        pdf.set_x(img_x)
                        pdf.image(path, w=img_w)
                        pdf.ln(8)
                        
                        try:
                            os.remove(path)
                        except:
                            pass
                
                pdf.output(pdf_path)
                
                status.update(label="✅ Análisis y PDF completados", state="complete", expanded=False)
                
                # Mostrar en Streamlit
                st.markdown(f'<div class="ia-result">{resultado}</div>', unsafe_allow_html=True)
                
                st.success(f"📄 El PDF se guardó correctamente en la carpeta 'analisis de informe' como: `{pdf_filename}`")
                
                # Botón para descargar el PDF generado
                with open(pdf_path, "rb") as f:
                    pdf_bytes = f.read()
                    
                st.download_button(
                    label="📥 Descargar PDF del Informe",
                    data=pdf_bytes,
                    file_name=pdf_filename,
                    mime="application/pdf",
                    type="primary"
                )
        except Exception as e:
            status.update(label="❌ Error en el análisis", state="error")
            st.error(f"Error: {str(e)}")


